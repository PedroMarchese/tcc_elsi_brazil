"""
Gera tabelas de regressão (OLS e logística) em formato APA 7ª edição.

Lê as tabelas-resumo geradas pelo script principal em tables/ e exporta
arquivos .xlsx formatados em tables/tabelas_apa/.

Regras APA aplicadas:
  - Apenas bordas horizontais: acima do cabeçalho, abaixo do cabeçalho e
    abaixo da última linha de dados (sem bordas verticais).
  - "Tabela N" em negrito (linha 1); título em itálico (linha 2).
  - Asteriscos de significância (* p<.05  ** p<.01  *** p<.001) concatenados
    ao valor de p.
  - p-valor sem zero à esquerda (ex.: ".032", "< .001").
  - IC 95% em coluna única no formato [inf, sup].
  - Nota em itálico abaixo da tabela, definindo abreviaturas e asteriscos.
"""

from pathlib import Path
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side, Alignment
from openpyxl.utils import get_column_letter

# ── Caminhos ──────────────────────────────────────────────────────────────────
SCRIPT_DIR = Path(__file__).resolve().parent
PROJECT_DIR = SCRIPT_DIR.parent
TABLES_DIR = PROJECT_DIR / "tables"
APA_DIR = TABLES_DIR / "tabelas_apa"
APA_DIR.mkdir(exist_ok=True)

# ── Rótulos legíveis ──────────────────────────────────────────────────────────
ROTULOS_DESFECHO = {
    "n_morbidades": "N.º de morbidades",
    "soma_grupo_0_neuro_oftalmo_musculoesqueleticas": "Soma — Gr. 0 (neuro/oftalmo/musculoesq.)",
    "soma_grupo_1_cardiorrespiratorias": "Soma — Gr. 1 (cardiorrespiratório)",
    "soma_grupo_2_outras": "Soma — Gr. 2 (outras)",
    "grupo_0_neuro_oftalmo_musculoesqueleticas": "Gr. 0 (neuro/oftalmo/musculoesq.)",
    "grupo_1_cardiorrespiratorias": "Gr. 1 (cardiorrespiratório)",
    "grupo_2_outras": "Gr. 2 (outras)",
    "morbidades_menor_que_2": "Morbidades < 2",
    "morbidades_igual_a_2": "Morbidades = 2",
    "morbidades_igual_a_3": "Morbidades = 3",
    "morbidades_igual_a_4": "Morbidades = 4",
    "morbidades_igual_a_5": "Morbidades = 5",
    "morbidades_igual_a_6": "Morbidades = 6",
    "morbidades_maior_ou_igual_a_7": "Morbidades ≥ 7",
}

ROTULOS_PREDITOR = {
    "SSWS": "SSWS (m/s)",
    "LRI": "LRI (%)",
}

# ── Bordas APA ────────────────────────────────────────────────────────────────
_LADO = Side(style="thin", color="000000")
_SEM  = Side(style=None)

BORDA_CABECALHO  = Border(top=_LADO, bottom=_LADO)   # acima + abaixo do header
BORDA_ULTIMA     = Border(bottom=_LADO)               # só abaixo da última linha


# ── Utilitários de formatação ─────────────────────────────────────────────────
def _sig(p):
    try:
        p = float(p)
    except (ValueError, TypeError):
        return ""
    if p < 0.001:
        return "***"
    if p < 0.01:
        return "**"
    if p < 0.05:
        return "*"
    return ""


def _fmt_p(p):
    """p-valor APA: sem zero à esquerda, mínimo "< .001"."""
    try:
        p = float(p)
    except (ValueError, TypeError):
        return str(p)
    if p < 0.001:
        return "< .001"
    partes = f"{p:.3f}".split(".")
    return f".{partes[1]}"


def _fmt_ic(inf, sup, casas=3):
    """Intervalo de confiança APA: [inf, sup]."""
    try:
        return f"[{float(inf):.{casas}f}, {float(sup):.{casas}f}]"
    except (ValueError, TypeError):
        return ""


def _celula(ws, row, col, valor, bold=False, italic=False, borda=None,
            alinhamento="left", fonte_size=11):
    cell = ws.cell(row=row, column=col, value=valor)
    cell.font = Font(bold=bold, italic=italic, name="Times New Roman", size=fonte_size)
    cell.alignment = Alignment(horizontal=alinhamento, wrap_text=True, vertical="center")
    if borda:
        cell.border = borda
    return cell


def _cabecalho(ws, row, colunas):
    for col_idx, nome in enumerate(colunas, start=1):
        cell = ws.cell(row=row, column=col_idx, value=nome)
        cell.font = Font(bold=True, name="Times New Roman", size=11)
        cell.border = BORDA_CABECALHO
        cell.alignment = Alignment(horizontal="center", wrap_text=True, vertical="center")


def _nota(ws, row, texto, n_colunas):
    cell = ws.cell(row=row, column=1, value=f"Note. {texto}")
    cell.font = Font(italic=True, name="Times New Roman", size=10)
    cell.alignment = Alignment(wrap_text=True, vertical="top")
    if n_colunas > 1:
        ws.merge_cells(start_row=row, start_column=1,
                       end_row=row, end_column=n_colunas)
    ws.row_dimensions[row].height = 60


def _ajustar_colunas(ws, larguras):
    for i, larg in enumerate(larguras, start=1):
        ws.column_dimensions[get_column_letter(i)].width = larg


def _ultima_linha_borda(ws, row, n_colunas):
    for col in range(1, n_colunas + 1):
        cell = ws.cell(row=row, column=col)
        cell.border = BORDA_ULTIMA


# ── OLS ───────────────────────────────────────────────────────────────────────
COLUNAS_OLS = [
    "Desfecho", "Preditor", "Modelo", "N",
    "β", "EP", "t", "p", "IC 95% (β)",
    "R²", "R² ajustado",
]
LARGURAS_OLS = [40, 11, 10, 7, 8, 8, 8, 10, 18, 7, 13]

NOTA_OLS = (
    "SSWS = velocidade de caminhada auto-selecionada (m/s); "
    "LRI = índice de reabilitação locomotor (%). "
    "β = coeficiente de regressão não-padronizado; EP = erro-padrão; "
    "IC 95% (β) = intervalo de confiança de 95% para β; "
    "R² aj. = coeficiente de determinação ajustado. "
    "O modelo Ajustado inclui como covariáveis: idade, sexo e escolaridade "
    "(C[escolaridade_modelo] — variáveis indicadoras; referência = menor nível de escolaridade). "
    "*p < .05. **p < .01. ***p < .001."
)


def _preparar_ols(df):
    linhas = []
    for _, row in df.iterrows():
        p_str = _fmt_p(row["p_value"]) + _sig(row["p_value"])
        linhas.append({
            "Desfecho":      ROTULOS_DESFECHO.get(row["desfecho"], row["desfecho"]),
            "Preditor":      ROTULOS_PREDITOR.get(row["preditor"], row["preditor"]),
            "Modelo":        "Ajustado" if str(row["ajustado"]).lower() == "true" else "Cru",
            "N":             int(row["n"]),
            "β":             round(float(row["coef"]), 4),
            "EP":            round(float(row["std_err"]), 4),
            "t":             round(float(row["stat"]), 3),
            "p":             p_str,
            "IC 95% (β)":   _fmt_ic(row["ic95_inf"], row["ic95_sup"], casas=4),
            "R²":            round(float(row["r_squared"]), 3),
            "R² ajustado":   round(float(row["r_squared_ajustado"]), 3),
        })
    return pd.DataFrame(linhas)


def salvar_ols_apa(df_raw, wave, num_tabela):
    df = _preparar_ols(df_raw)
    nc = len(COLUNAS_OLS)

    wb = Workbook()
    ws = wb.active
    ws.title = f"OLS Wave {wave}"
    ws.sheet_view.showGridLines = False

    # Linha 1 — número da tabela
    _celula(ws, 1, 1, f"Tabela {num_tabela}", bold=True)
    ws.row_dimensions[1].height = 16

    # Linha 2 — título
    titulo = (
        f"Regressão Linear (OLS): Associação de SSWS e LRI com Desfechos "
        f"de Morbidade — Wave {wave}"
    )
    _celula(ws, 2, 1, titulo, italic=True)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=nc)
    ws.row_dimensions[2].height = 28

    # Linha 3 — cabeçalho com bordas APA
    _cabecalho(ws, 3, COLUNAS_OLS)
    ws.row_dimensions[3].height = 28

    # Linhas de dados
    ALIN_DIREITA = {"N", "β", "EP", "t", "R²", "R² ajustado"}
    ALIN_CENTRO  = {"p", "IC 95% (β)", "Modelo"}
    for i, (_, row) in enumerate(df.iterrows(), start=4):
        for col_idx, col_nome in enumerate(COLUNAS_OLS, start=1):
            if col_nome in ALIN_DIREITA:
                alin = "right"
            elif col_nome in ALIN_CENTRO:
                alin = "center"
            else:
                alin = "left"
            _celula(ws, i, col_idx, row[col_nome], alinhamento=alin)
        ws.row_dimensions[i].height = 14

    ultima = 3 + len(df)
    _ultima_linha_borda(ws, ultima, nc)
    _nota(ws, ultima + 2, NOTA_OLS, nc)
    _ajustar_colunas(ws, LARGURAS_OLS)

    caminho = APA_DIR / f"ols_w{wave}_apa.xlsx"
    wb.save(caminho)
    print(f"[APA] {caminho.name}")


# ── Logística ─────────────────────────────────────────────────────────────────
COLUNAS_LOG = [
    "Desfecho", "Preditor", "Modelo", "N",
    "OR", "EP (β)", "z", "p", "IC 95% (OR)",
    "R² pseudo (McFadden)",
]
LARGURAS_LOG = [40, 11, 10, 7, 8, 8, 8, 10, 18, 22]

NOTA_LOG = (
    "SSWS = velocidade de caminhada auto-selecionada (m/s); "
    "LRI = índice de reabilitação locomotor (%). "
    "OR = razão de chances (odds ratio); EP (β) = erro-padrão do coeficiente logístico; "
    "IC 95% (OR) = intervalo de confiança de 95% para o OR; "
    "R² pseudo = R² de McFadden (pseudo-R² de ajuste global do modelo). "
    "O modelo Ajustado inclui como covariáveis: idade, sexo e escolaridade "
    "(C[escolaridade_modelo] — variáveis indicadoras; referência = menor nível de escolaridade). "
    "*p < .05. **p < .01. ***p < .001."
)


def _preparar_log(df):
    linhas = []
    for _, row in df.iterrows():
        p_str = _fmt_p(row["p_value"]) + _sig(row["p_value"])
        linhas.append({
            "Desfecho":             ROTULOS_DESFECHO.get(row["desfecho"], row["desfecho"]),
            "Preditor":             ROTULOS_PREDITOR.get(row["preditor"], row["preditor"]),
            "Modelo":               "Ajustado" if str(row["ajustado"]).lower() == "true" else "Cru",
            "N":                    int(row["n"]),
            "OR":                   round(float(row["odds_ratio"]), 4),
            "EP (β)":               round(float(row["std_err"]), 4),
            "z":                    round(float(row["stat"]), 3),
            "p":                    p_str,
            "IC 95% (OR)":          _fmt_ic(row["or_ic95_inf"], row["or_ic95_sup"], casas=4),
            "R² pseudo (McFadden)": round(float(row["pseudo_r_squared"]), 3),
        })
    return pd.DataFrame(linhas)


def salvar_log_apa(df_raw, wave, num_tabela):
    df = _preparar_log(df_raw)
    nc = len(COLUNAS_LOG)

    wb = Workbook()
    ws = wb.active
    ws.title = f"Logística Wave {wave}"
    ws.sheet_view.showGridLines = False

    _celula(ws, 1, 1, f"Tabela {num_tabela}", bold=True)
    ws.row_dimensions[1].height = 16

    titulo = (
        f"Regressão Logística: Associação de SSWS e LRI com Desfechos "
        f"Dicotômicos de Morbidade — Wave {wave}"
    )
    _celula(ws, 2, 1, titulo, italic=True)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=nc)
    ws.row_dimensions[2].height = 28

    _cabecalho(ws, 3, COLUNAS_LOG)
    ws.row_dimensions[3].height = 28

    ALIN_DIREITA = {"N", "OR", "EP (β)", "z", "R² pseudo (McFadden)"}
    ALIN_CENTRO  = {"p", "IC 95% (OR)", "Modelo"}
    for i, (_, row) in enumerate(df.iterrows(), start=4):
        for col_idx, col_nome in enumerate(COLUNAS_LOG, start=1):
            if col_nome in ALIN_DIREITA:
                alin = "right"
            elif col_nome in ALIN_CENTRO:
                alin = "center"
            else:
                alin = "left"
            _celula(ws, i, col_idx, row[col_nome], alinhamento=alin)
        ws.row_dimensions[i].height = 14

    ultima = 3 + len(df)
    _ultima_linha_borda(ws, ultima, nc)
    _nota(ws, ultima + 2, NOTA_LOG, nc)
    _ajustar_colunas(ws, LARGURAS_LOG)

    caminho = APA_DIR / f"logistica_w{wave}_apa.xlsx"
    wb.save(caminho)
    print(f"[APA] {caminho.name}")


# ── Main ──────────────────────────────────────────────────────────────────────
def main():
    num = 1
    for wave in [1, 2]:
        ols_path = TABLES_DIR / f"regressoes_ols_w{wave}.csv"
        log_path = TABLES_DIR / f"regressoes_logisticas_w{wave}.csv"

        if ols_path.exists():
            salvar_ols_apa(pd.read_csv(ols_path), wave, num)
            num += 1
        else:
            print(f"[AVISO] não encontrado: {ols_path.name}")

        if log_path.exists():
            salvar_log_apa(pd.read_csv(log_path), wave, num)
            num += 1
        else:
            print(f"[AVISO] não encontrado: {log_path.name}")

    print(f"\nTabelas APA geradas em: {APA_DIR}")


if __name__ == "__main__":
    main()
